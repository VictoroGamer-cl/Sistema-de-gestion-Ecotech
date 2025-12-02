import mysql.connector
from KEYS import key
from EcoTech import EcoTech
from Departamento import Departamento
from Administrador import Administrador
from Usuarios import Usuario
from HoraTrabajada import HoraTrabajada
from Proyecto import Proyecto
from Informe import Informe
import hashlib
from config_emb import get_db_config
import platform
from datetime import datetime
import os
import requests

class DAO:

    def cifrar_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()
    
    def __init__(self):
        self.config = get_db_config()
        self.conn = None
        self.cursor = None
        self.conectar()

    def conectar(self):
        try:
            from config_db import DB_CONFIG
            self.conn = mysql.connector.connect(**DB_CONFIG)
            print("Conexión a Railway exitosa")
        except mysql.connector.Error as err:
            print(f"Error al conectar a MySQL: {err}")

    def desconectar(self):
        if self.conn and self.conn.is_connected():
            self.cursor.close()
            self.conn.close()
        print("La conexion a MySQL se ha cerrado")
        
    def registrar_empleado(self, empleado):
        try:
            self.cursor = self.conn.cursor()
            sql = """INSERT INTO empleados 
                 (idEmpleados, nombreEmpleado, direccionEmpleado, telefonoEmpleado, 
                  mailEmpleado, fechaInicioContrato, salarioEmpleado, idDepartamento, idAdministrador) 
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            valores = (
                empleado.get_id(),
                empleado.get_nombre(),
                empleado.get_direccion(),
                empleado.get_telefono(),
                empleado.get_mail(),
                empleado.get_fecheIC(),
                empleado.get_salario(),
                empleado.get_id_departamento(),
                empleado.get_id_administrador()
            )
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"Empleado {empleado.get_nombre()} registrado exitosamente.")
            try:
                return self.cursor.lastrowid
            except Exception:
                return None
        except Exception as e:
            print(f"[DAO] Error en registrar_empleado: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            return None
    
    def obtener_empleados(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM empleados"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()

            empleados = []
            for fila in resultados:
                try:
                    empleado = EcoTech(
                        fila[0], fila[1], fila[2], fila[3], fila[4], fila[5], fila[6], fila[7]
                    )
                except Exception:
                   
                    continue
                empleados.append(empleado)

            return empleados
        except Exception as e:
            print(f"[DAO] Error en obtener_empleados: {e}")
            return []
        
    def actualizar_empleado(self, empleado):
        try:
            self.cursor = self.conn.cursor()
            sql = "UPDATE empleados SET nombreEmpleado = %s, direccionEmpleado = %s, telefonoEmpleado = %s, mailEmpleado = %s, fechaInicioContrato = %s, salarioEmpleado = %s WHERE idEmpleadoS = %s"
            valores = (
                empleado.get_nombre(),
                empleado.get_direccion(),
                empleado.get_telefono(),
                empleado.get_mail(),
                empleado.get_fecheIC(),
                empleado.get_salario(),
                empleado.get_id()
            )
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"Empleado {empleado.get_nombre()} actualizado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en actualizar_empleado: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def eliminar_empleado(self, id_empleado):
        try:
            self.cursor = self.conn.cursor()
            
            try:
                self.conn.start_transaction()
            except Exception:
                pass
            dependent_deletes = [
                ("DELETE FROM usuarios WHERE idEmpleado = %s", (id_empleado,)),
                ("DELETE FROM horastrabajadas WHERE idEmpleado = %s", (id_empleado,)),
                
                ("DELETE FROM empleados WHERE idEmpleado = %s", (id_empleado,)),
                ("DELETE FROM usuarios WHERE idEmpleadoS = %s", (id_empleado,)),
                ("DELETE FROM horastrabajadas WHERE idEmpleadoS = %s", (id_empleado,)),
            ]

            for sql, params in dependent_deletes:
                try:
                    self.cursor.execute(sql, params)
                except Exception:
    
                    pass

            
            empleado_deletes = [
                ("DELETE FROM empleados WHERE idEmpleadoS = %s", (id_empleado,)),
                ("DELETE FROM empleados WHERE idEmpleados = %s", (id_empleado,)),
                ("DELETE FROM empleados WHERE idEmpleado = %s", (id_empleado,)),
            ]

            eliminado_ok = False
            for sql, params in empleado_deletes:
                try:
                    self.cursor.execute(sql, params)
                    if getattr(self.cursor, 'rowcount', 0) != 0:
                        eliminado_ok = True
                except Exception:
                    pass

            
            try:
                self.conn.commit()
            except Exception as e:
                try:
                    self.conn.rollback()
                except Exception:
                    pass
                self.last_error = str(e)
                return False

            if not eliminado_ok:
                self.last_error = f"No se encontró el empleado con id {id_empleado} o no se pudo eliminar." 
                return False

            self.last_error = None
            return True

        except Exception as e:
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            self.last_error = str(e)
            return False

    def registrar_departamento(self, departamento):
        try:
            self.cursor = self.conn.cursor()
            sql = "INSERT INTO departamento (Gerente, nombreDepartamento, idGerente) VALUES (%s, %s, %s)"
            valores = (departamento.get_nombre_gerente(), departamento.get_nombre(), departamento.get_id_gerente())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"departamento {departamento.get_nombre()} registrado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en registrar_departamento: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def obtener_departamentos(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM departamento"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()

            departamentos = []
            for fila in resultados:
                try:
                    depto = Departamento(fila[0], fila[2], fila[3], fila[1]) 
                except Exception:
                    continue
                departamentos.append(depto)

            return departamentos
        except Exception as e:
            print(f"[DAO] Error en obtener_departamentos: {e}")
            return []

    def asignar_empleado_departamento(self, id_empleado, id_departamento):
        try:
            self.cursor = self.conn.cursor()
            sql = "UPDATE empleados SET idDepartamento = %s WHERE idEmpleados = %s"
            self.cursor.execute(sql, (id_departamento, id_empleado))
            self.conn.commit()
            print(f"Empleado {id_empleado} asignado al departamento {id_departamento}.")
        except Exception as e:
            print(f"[DAO] Error en asignar_empleado_departamento: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
    
    def eliminar_departamento(self, id_departamento):
        try:
            self.cursor = self.conn.cursor()
            sql = "DELETE FROM departamento WHERE idDepartamento = %s"
            self.cursor.execute(sql, (id_departamento,))
            self.conn.commit()
            print(f"Departamento con ID {id_departamento} eliminado de la base de datos.")
        except Exception as e:
            print(f"[DAO] Error en eliminar_departamento: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def registrar_administrador(self, administrador):
        try:
            self.cursor = self.conn.cursor()
            password_cifrada = self.cifrar_password(administrador.get_password())
            sql = "INSERT INTO administradores (username, password, email, nivel_acceso) VALUES (%s, %s, %s, %s)"
            valores = (administrador.get_username(), password_cifrada, administrador.get_email(), administrador.get_nivel_acceso())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"Administrador {administrador.get_username()} registrado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en registrar_administrador: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def obtener_administradores(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM administradores"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()
            administradores = []
            for fila in resultados:
                try:
                    admin = Administrador(fila[0], fila[1], fila[2], fila[3], fila[4])
                except Exception:
                    continue
                administradores.append(admin)
            return administradores
        except Exception as e:
            print(f"[DAO] Error en obtener_administradores: {e}")
            return []
    
    def cifrar_password(self, password):
        try:
            import hashlib
            return hashlib.sha256(password.encode()).hexdigest()
        except Exception as e:
            print(f"Error cifrando password: {e}")
            return None
        
    def login_administrador(self, username, password):
        try:
            if not hasattr(self, 'conn') or not self.conn:
                print("[DAO] No hay conexión a la base de datos")
                return None
            self.cursor = self.conn.cursor()
            password_cifrada = self.cifrar_password(password)
            sql = "SELECT * FROM administradores WHERE username = %s AND password = %s"
            self.cursor.execute(sql, (username, password_cifrada))
            resultado = self.cursor.fetchone()
            if resultado:
                return Administrador(resultado[0], resultado[1], resultado[2], resultado[3], resultado[4])
            return None
        except Exception as e:
            print(f"[DAO] Error en login_administrador: {e}")
            return None
    
    def registrar_usuario(self, usuario):
        try:
            self.cursor = self.conn.cursor()
            password_cifrada = self.cifrar_password(usuario.get_contraseña())
            sql = "INSERT INTO usuarios (username, contraseña, modulos, idEmpleado) VALUES (%s, %s, %s, %s)"
            valores = (usuario.get_username(), password_cifrada, usuario.get_modulos(), usuario.get_id_empleado())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"Usuario {usuario.get_username()} registrado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en registrar_usuario: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def obtener_usuarios(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM usuarios"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()
            usuarios = []
            for fila in resultados:
                try:
                    usuario = Usuario(fila[0], fila[1], fila[2], fila[3], fila[4])
                except Exception:
                    continue
                usuarios.append(usuario)
            return usuarios
        except Exception as e:
            print(f"[DAO] Error en obtener_usuarios: {e}")
            return []

    def login_usuario(self, username, password):
        try:
            if not hasattr(self, 'conn') or not self.conn:
                print("[DAO] No hay conexión a la base de datos")
                return None
                
            self.cursor = self.conn.cursor()
            password_cifrada = self.cifrar_password(password)  # ← ESTA LÍNEA ES CLAVE
            
            sql = "SELECT * FROM usuarios WHERE username = %s AND contraseña = %s"
            self.cursor.execute(sql, (username, password_cifrada))
            resultado = self.cursor.fetchone()
            
            if resultado:
                return Usuario(
                    id_usuario=resultado[0],
                    username=resultado[1],
                    contraseña=resultado[2],
                    modulos=resultado[3],
                    id_empleado=resultado[4]
                )
            return None
        except Exception as e:
            print(f"[DAO] Error en login_usuario: {e}")
            return None

    def registrar_hora_trabajada(self, hora_trabajada):
        try:
            self.cursor = self.conn.cursor()
            sql = "INSERT INTO horastrabajadas (fecha, horas_trabajadas, descripcion_tareas, idEmpleado) VALUES (%s, %s, %s, %s)"
            valores = (hora_trabajada.get_fecha(), hora_trabajada.get_horas_trabajadas(), 
                       hora_trabajada.get_descripcion_tareas(), hora_trabajada.get_id_empleado())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print("Horas registradas exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en registrar_hora_trabajada: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def obtener_horas_empleado(self, id_empleado):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM horastrabajadas WHERE idEmpleado = %s ORDER BY fecha DESC"
            self.cursor.execute(sql, (id_empleado,))
            resultados = self.cursor.fetchall()
            horas = []
            for fila in resultados:
                try:
                    hora = HoraTrabajada(fila[0], fila[1], fila[2], fila[3], fila[4])
                except Exception:
                    continue
                horas.append(hora)
            return horas
        except Exception as e:
            print(f"[DAO] Error en obtener_horas_empleado: {e}")
            return []

    def registrar_proyecto(self, proyecto):
        try:
            self.cursor = self.conn.cursor()
            sql = "INSERT INTO proyectos (nombre_proyecto, descripcion, fecha_inicio, idDepartamento) VALUES (%s, %s, %s, %s)"
            valores = (proyecto.get_nombre_proyecto(), proyecto.get_descripcion(), 
                       proyecto.get_fecha_inicio(), proyecto.get_id_departamento())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"proyectos {proyecto.get_nombre_proyecto()} registrado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en registrar_proyecto: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def obtener_proyectos(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM proyectos"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()
            proyectos = []
            for fila in resultados:
                try:
                    proyecto = Proyecto(fila[0], fila[1], fila[2], fila[3], fila[4])
                except Exception:
                    continue
                proyectos.append(proyecto)
            return proyectos
        except Exception as e:
            print(f"[DAO] Error en obtener_proyectos: {e}")
            return []

    def actualizar_proyecto(self, proyecto):
        try:
            self.cursor = self.conn.cursor()
            sql = "UPDATE proyectos SET nombre_proyecto = %s, descripcion = %s, fecha_inicio = %s, idDepartamento = %s WHERE idProyecto = %s"
            valores = (proyecto.get_nombre_proyecto(), proyecto.get_descripcion(), 
                       proyecto.get_fecha_inicio(), proyecto.get_id_departamento(), proyecto.get_id())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            print(f"Proyecto {proyecto.get_nombre_proyecto()} actualizado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en actualizar_proyecto: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def eliminar_proyecto(self, id_proyecto):
        try:
            self.cursor = self.conn.cursor()
            sql = "DELETE FROM proyectos WHERE idProyecto = %s"
            self.cursor.execute(sql, (id_proyecto,))
            self.conn.commit()
            print(f"Proyecto con ID {id_proyecto} eliminado exitosamente.")
        except Exception as e:
            print(f"[DAO] Error en eliminar_proyecto: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass

    def generar_reporte_empleados_pdf(self, empleados, nombre_archivo):
        
        try:
            print("🌐 Generando PDF ")
            
            
            html_content = self._crear_html_empleados(empleados)
            
            
            payload = {
                "source": html_content,
                "format": "A4",
                "margin": "10mm",
                "landscape": False
            }
            
            
            response = requests.post(
                "https://api.pdfshift.io/v3/convert/pdf",
                json=payload,
                auth=('api', 'sk_8a1b3c5d7e9f0a1b3c5d7e9f0a1b3c5d'),  
                timeout=30
            )
            
            
            if response.status_code == 200:
                carpeta_informes = self.crear_carpeta_informes()
                ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
                
                
                with open(ruta_pdf, 'wb') as f:
                    f.write(response.content)
                
                print(f"✅ PDF generado exitosamente via API: {ruta_pdf}")
                return ruta_pdf
            else:
                print(f"❌ Error API PDFShift: {response.status_code} - {response.text}")
                
                print("🔄 Intentando generación local como fallback...")
                return self._generar_reporte_empleados_pdf_local(empleados, nombre_archivo)
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Error de conexión API PDF: {e}")
            
            return self._generar_reporte_empleados_pdf_local(empleados, nombre_archivo)
        except Exception as e:
            print(f"❌ Error inesperado: {e}")
            return self._generar_reporte_empleados_pdf_local(empleados, nombre_archivo)
        
    def _crear_html_empleados(self, empleados):
        
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        total_salarios = sum(emp.get_salario() for emp in empleados)
        promedio_salario = total_salarios / len(empleados) if empleados else 0
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Informe de Empleados - EcoTech</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px; 
                    color: #333;
                }}
                .header {{ 
                    text-align: center; 
                    color: #2E8B57; 
                    border-bottom: 2px solid #2E8B57;
                    padding-bottom: 10px;
                }}
                .info {{ 
                    margin: 15px 0; 
                    font-size: 14px;
                }}
                table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                    margin-top: 20px;
                    font-size: 12px;
                }}
                th, td {{ 
                    border: 1px solid #ddd; 
                    padding: 10px; 
                    text-align: left;
                }}
                th {{ 
                    background-color: #2E8B57; 
                    color: white; 
                    font-weight: bold;
                }}
                tr:nth-child(even) {{ 
                    background-color: #f8f9fa; 
                }}
                .stats {{ 
                    margin-top: 20px; 
                    padding: 15px;
                    background-color: #e9f7ef;
                    border-left: 4px solid #2E8B57;
                }}
                .footer {{
                    margin-top: 30px;
                    text-align: center;
                    font-size: 12px;
                    color: #666;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>INFORME DE EMPLEADOS - ECOTECH</h1>
            </div>
            
            <div class="info">
                <p><strong>Fecha de generación:</strong> {fecha_actual}</p>
                <p><strong>Total empleados:</strong> {len(empleados)}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Nombre</th>
                        <th>Email</th>
                        <th>Teléfono</th>
                        <th>Salario</th>
                        <th>Fecha Ingreso</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for emp in empleados:
            html += f"""
                    <tr>
                        <td>{emp.get_id()}</td>
                        <td>{emp.get_nombre()}</td>
                        <td>{emp.get_mail()}</td>
                        <td>{emp.get_telefono()}</td>
                        <td>${emp.get_salario():,}</td>
                        <td>{emp.get_fecheIC()}</td>
                    </tr>
            """
        
        html += f"""
                </tbody>
            </table>
            
            <div class="stats">
                <p><strong>Estadísticas:</strong></p>
                <p>• Salario total: ${total_salarios:,}</p>
                <p>• Salario promedio: ${promedio_salario:,.2f}</p>
            </div>
            
            <div class="footer">
                <p>Generado automáticamente por Sistema EcoTech | {fecha_actual}</p>
            </div>
        </body>
        </html>
        """
        return html
    
    def _generar_reporte_empleados_pdf_local(self, empleados, nombre_archivo):
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            carpeta_informes = self.crear_carpeta_informes()
            ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
            
            doc = SimpleDocTemplate(ruta_pdf, pagesize=A4, topMargin=1*inch)
            elements = []
            
            styles = getSampleStyleSheet()
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  
            )
            
            titulo = Paragraph("INFORME DE EMPLEADOS - ECOTECH", titulo_style)
            elements.append(titulo)
            
            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
            info_text = f"<b>Fecha de generación:</b> {fecha_actual} | <b>Total empleados:</b> {len(empleados)}"
            info_paragraph = Paragraph(info_text, styles["Normal"])
            elements.append(info_paragraph)
            elements.append(Spacer(1, 0.2*inch))
            
            datos = [['ID', 'Nombre', 'Email', 'Teléfono', 'Salario', 'Fecha Ingreso']]
            
            for empleado in empleados:
                datos.append([
                    str(empleado.get_id()),
                    empleado.get_nombre(),
                    empleado.get_mail(),
                    str(empleado.get_telefono()),
                    f"${empleado.get_salario():,}",
                    empleado.get_fecheIC()
                ])
            
            tabla = Table(datos, colWidths=[0.5*inch, 1.5*inch, 1.8*inch, 1*inch, 1*inch, 1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),  
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')), 
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))
            
            elements.append(tabla)
            
            elements.append(Spacer(1, 0.3*inch))
            total_salarios = sum(emp.get_salario() for emp in empleados)
            promedio_salario = total_salarios / len(empleados) if empleados else 0
            
            stats_text = f"<b>Estadísticas:</b> Salario total: ${total_salarios:,} | Promedio: ${promedio_salario:,.2f}"
            stats_paragraph = Paragraph(stats_text, styles["Normal"])
            elements.append(stats_paragraph)
            
            doc.build(elements)
            
            print("📄 PDF generado localmente (fallback)")
            return ruta_pdf
            
        except Exception as e:
            print(f"❌ Error en generación local: {e}")
            return f"Error: {str(e)}"

    def generar_reporte_empleados_excel(self, empleados, nombre_archivo):
        
        try:
            print("🌐 Generando Excel via API APYHub...")
            
            
            datos_tabla = {
                "title": f"Informe Empleados - {datetime.now().strftime('%Y-%m-%d')}",
                "headers": ["ID", "Nombre", "Dirección", "Teléfono", "Email", "Fecha Ingreso", "Salario"],
                "rows": []
            }
            
            for emp in empleados:
                datos_tabla["rows"].append([
                    emp.get_id(),
                    emp.get_nombre(),
                    emp.get_direccion() or "",
                    emp.get_telefono(),
                    emp.get_mail(),
                    emp.get_fecheIC(),
                    emp.get_salario()
                ])
            
            # ✅ CONSUMO DE API EXTERNA
            response = requests.post(
                "https://api.apyhub.com/generate/excel/table",
                json=datos_tabla,
                headers={"apy-token": "APY0xXxXxXxXxXxXxXxXxXxXxXxXxXxXxXxXxXxXxXx"},  # API key de ejemplo
                timeout=30
            )
            
            if response.status_code == 200:
                carpeta_informes = self.crear_carpeta_informes()
                ruta_excel = os.path.join(carpeta_informes, f"{nombre_archivo}.xlsx")
                
                with open(ruta_excel, 'wb') as f:
                    f.write(response.content)
                
                print(f"✅ Excel generado exitosamente via API: {ruta_excel}")
                return ruta_excel
            else:
                print(f"❌ Error API Excel: {response.status_code}")
                # ✅ FALLBACK a generación local
                print("🔄 Intentando generación local como fallback...")
                return self._generar_reporte_empleados_excel_local(empleados, nombre_archivo)
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Error de conexión API Excel: {e}")
            return self._generar_reporte_empleados_excel_local(empleados, nombre_archivo)
        except Exception as e:
            print(f"❌ Error inesperado: {e}")
            return self._generar_reporte_empleados_excel_local(empleados, nombre_archivo)

    def _generar_reporte_empleados_excel_local(self, empleados, nombre_archivo):
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            carpeta_informes = self.crear_carpeta_informes()
            ruta_excel = os.path.join(carpeta_informes, f"{nombre_archivo}.xlsx")
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Empleados"
            
            header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            
            encabezados = ['ID', 'Nombre', 'Dirección', 'Teléfono', 'Email', 'Fecha Ingreso', 'Salario']
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col, value=encabezado)
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = center_align
            
            for fila, empleado in enumerate(empleados, 2):
                ws.cell(row=fila, column=1, value=empleado.get_id())
                ws.cell(row=fila, column=2, value=empleado.get_nombre())
                ws.cell(row=fila, column=3, value=empleado.get_direccion())
                ws.cell(row=fila, column=4, value=empleado.get_telefono())
                ws.cell(row=fila, column=5, value=empleado.get_mail())
                ws.cell(row=fila, column=6, value=empleado.get_fecheIC())
                ws.cell(row=fila, column=7, value=empleado.get_salario())
            
            for col in range(1, len(encabezados) + 1):
                col_letra = get_column_letter(col)
                if col == 2: 
                    ws.column_dimensions[col_letra].width = 25
                elif col == 3:  
                    ws.column_dimensions[col_letra].width = 30
                elif col == 5: 
                    ws.column_dimensions[col_letra].width = 25
                else:
                    ws.column_dimensions[col_letra].width = 15
            
            ws_stats = wb.create_sheet("Estadísticas")
            
            total_empleados = len(empleados)
            total_salarios = sum(emp.get_salario() for emp in empleados)
            promedio_salario = total_salarios / total_empleados if total_empleados else 0
            salario_max = max(emp.get_salario() for emp in empleados) if empleados else 0
            salario_min = min(emp.get_salario() for emp in empleados) if empleados else 0
            
            stats_data = [
                ["Estadísticas de Empleados", ""],
                ["Total de empleados", total_empleados],
                ["Salario total", f"${total_salarios:,}"],
                ["Salario promedio", f"${promedio_salario:,.2f}"],
                ["Salario máximo", f"${salario_max:,}"],
                ["Salario mínimo", f"${salario_min:,}"],
                ["", ""],
                ["Generado el", datetime.now().strftime("%d/%m/%Y %H:%M")]
            ]
            
            for fila, (concepto, valor) in enumerate(stats_data, 1):
                ws_stats.cell(row=fila, column=1, value=concepto).font = Font(bold=True)
                ws_stats.cell(row=fila, column=2, value=valor)
            
            wb.save(ruta_excel)
            
            print("📊 Excel generado localmente (fallback)")
            return ruta_excel
            
        except Exception as e:
            print(f"❌ Error en generación local Excel: {e}")
            return f"Error: {str(e)}"

    def generar_reporte_proyectos_pdf(self, proyectos, nombre_archivo):
        
        try:
            print("🌐 Generando PDF")
            
            html_content = self._crear_html_proyectos(proyectos)
            
            payload = {
                "source": html_content,
                "format": "A4",
                "margin": "10mm"
            }
            
            response = requests.post(
                "https://api.pdfshift.io/v3/convert/pdf",
                json=payload,
                auth=('api', 'sk_8a1b3c5d7e9f0a1b3c5d7e9f0a1b3c5d'),  # Misma key que empleados
                timeout=30
            )
            
            if response.status_code == 200:
                carpeta_informes = self.crear_carpeta_informes()
                ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
                
                with open(ruta_pdf, 'wb') as f:
                    f.write(response.content)
                
                print(f"✅ PDF de proyectos generado via API: {ruta_pdf}")
                return ruta_pdf
            else:
                print(f"❌ Error API PDF proyectos: {response.status_code}")
                return self._generar_reporte_proyectos_pdf_local(proyectos, nombre_archivo)
                
        except Exception as e:
            print(f"❌ Error generando PDF proyectos: {e}")
            return self._generar_reporte_proyectos_pdf_local(proyectos, nombre_archivo)

    def _crear_html_proyectos(self, proyectos):
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        proyectos_sin_depto = sum(1 for p in proyectos if not p.get_id_departamento())
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Informe de Proyectos - EcoTech</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px; 
                    color: #333;
                }}
                .header {{ 
                    text-align: center; 
                    color: #2E8B57; 
                    border-bottom: 2px solid #2E8B57;
                    padding-bottom: 10px;
                }}
                .info {{ 
                    margin: 15px 0; 
                    font-size: 14px;
                }}
                table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                    margin-top: 20px;
                    font-size: 12px;
                }}
                th, td {{ 
                    border: 1px solid #ddd; 
                    padding: 10px; 
                    text-align: left;
                }}
                th {{ 
                    background-color: #2E8B57; 
                    color: white; 
                    font-weight: bold;
                }}
                tr:nth-child(even) {{ 
                    background-color: #f8f9fa; 
                }}
                .stats {{ 
                    margin-top: 20px; 
                    padding: 15px;
                    background-color: #e9f7ef;
                    border-left: 4px solid #2E8B57;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>INFORME DE PROYECTOS - ECOTECH</h1>
            </div>
            
            <div class="info">
                <p><strong>Fecha de generación:</strong> {fecha_actual}</p>
                <p><strong>Total proyectos:</strong> {len(proyectos)}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Nombre</th>
                        <th>Descripción</th>
                        <th>Fecha Inicio</th>
                        <th>Depto ID</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for proyecto in proyectos:
            descripcion = proyecto.get_descripcion()
            if descripcion and len(descripcion) > 50:
                descripcion = descripcion[:50] + "..."
            elif not descripcion:
                descripcion = "Sin descripción"
            
            html += f"""
                    <tr>
                        <td>{proyecto.get_id()}</td>
                        <td>{proyecto.get_nombre_proyecto() or 'Sin nombre'}</td>
                        <td>{descripcion}</td>
                        <td>{proyecto.get_fecha_inicio() or 'N/A'}</td>
                        <td>{proyecto.get_id_departamento() or 'N/A'}</td>
                    </tr>
            """
        
        html += f"""
                </tbody>
            </table>
            
            <div class="stats">
                <p><strong>Estadísticas:</strong> Proyectos sin departamento: {proyectos_sin_depto}</p>
            </div>
        </body>
        </html>
        """
        return html

    def _generar_reporte_proyectos_pdf_local(self, proyectos, nombre_archivo):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            carpeta_informes = self.crear_carpeta_informes()
            ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
            
            doc = SimpleDocTemplate(ruta_pdf, pagesize=A4, topMargin=1*inch)
            elements = []
            
            styles = getSampleStyleSheet()
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  
            )
            
            titulo = Paragraph("INFORME DE PROYECTOS - ECOTECH", titulo_style)
            elements.append(titulo)

            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
            info_text = f"<b>Fecha de generación:</b> {fecha_actual} | <b>Total proyectos:</b> {len(proyectos)}"
            info_paragraph = Paragraph(info_text, styles["Normal"])
            elements.append(info_paragraph)
            elements.append(Spacer(1, 0.2*inch))

            datos = [['ID', 'Nombre', 'Descripción', 'Fecha Inicio', 'Depto ID']]
            
            for proyecto in proyectos:
                descripcion = proyecto.get_descripcion()
                if descripcion and len(descripcion) > 50:
                    descripcion = descripcion[:50] + "..."
                elif not descripcion:
                    descripcion = "Sin descripción"
                
                try:
                    id_proyecto = str(proyecto.get_id())
                    nombre = proyecto.get_nombre_proyecto() or "Sin nombre"
                    fecha_inicio = proyecto.get_fecha_inicio() or "N/A"
                    id_depto = str(proyecto.get_id_departamento()) if proyecto.get_id_departamento() else "N/A"
                    
                    datos.append([id_proyecto, nombre, descripcion, fecha_inicio, id_depto])
                    
                except Exception as e:
                    print(f"Error procesando proyecto: {e}")
                    continue

            tabla = Table(datos, colWidths=[0.6*inch, 1.8*inch, 3.0*inch, 1.0*inch, 0.8*inch])
            
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))
            
            elements.append(tabla)
            
            elements.append(Spacer(1, 0.3*inch))
            proyectos_sin_depto = sum(1 for p in proyectos if not p.get_id_departamento())
            
            stats_text = f"<b>Estadísticas:</b> Proyectos sin departamento: {proyectos_sin_depto}"
            stats_paragraph = Paragraph(stats_text, styles["Normal"])
            elements.append(stats_paragraph)
            
            doc.build(elements)
            
            print(f"PDF de proyectos generado localmente (fallback): {ruta_pdf}")
            return ruta_pdf
            
        except Exception as e:
            print(f"Error generando PDF de proyectos local: {e}")
            return f"Error: {str(e)}"

    def generar_reporte_departamentos_pdf(self, departamentos, nombre_archivo):
        
        try:
            print("🌐 Generando PDF")
            
            html_content = self._crear_html_departamentos(departamentos)
            
            payload = {
                "source": html_content,
                "format": "A4",
                "margin": "10mm"
            }
            
            response = requests.post(
                "https://api.pdfshift.io/v3/convert/pdf",
                json=payload,
                auth=('api', 'sk_8a1b3c5d7e9f0a1b3c5d7e9f0a1b3c5d'),  # Misma key
                timeout=30
            )
            
            if response.status_code == 200:
                carpeta_informes = self.crear_carpeta_informes()
                ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
                
                with open(ruta_pdf, 'wb') as f:
                    f.write(response.content)
                
                print(f"✅ PDF de departamentos generado via API: {ruta_pdf}")
                return ruta_pdf
            else:
                print(f"❌ Error API PDF departamentos: {response.status_code}")
                return self._generar_reporte_departamentos_pdf_local(departamentos, nombre_archivo)
                
        except Exception as e:
            print(f"❌ Error generando PDF departamentos: {e}")
            return self._generar_reporte_departamentos_pdf_local(departamentos, nombre_archivo)

    def _crear_html_departamentos(self, departamentos):
        
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        empleados = self.obtener_empleados()
        empleados_por_depto = {}
        for emp in empleados:
            depto_id = emp.get_id_departamento()
            if depto_id:
                if depto_id not in empleados_por_depto:
                    empleados_por_depto[depto_id] = 0
                empleados_por_depto[depto_id] += 1
        
        total_empleados = sum(empleados_por_depto.values())
        deptos_sin_empleados = sum(1 for depto in departamentos if empleados_por_depto.get(depto.get_id(), 0) == 0)
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Informe de Departamentos - EcoTech</title>
            <style>
                body {{ 
                    font-family: Arial, sans-serif; 
                    margin: 20px; 
                    color: #333;
                }}
                .header {{ 
                    text-align: center; 
                    color: #2E8B57; 
                    border-bottom: 2px solid #2E8B57;
                    padding-bottom: 10px;
                }}
                .info {{ 
                    margin: 15px 0; 
                    font-size: 14px;
                }}
                table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                    margin-top: 20px;
                    font-size: 12px;
                }}
                th, td {{ 
                    border: 1px solid #ddd; 
                    padding: 10px; 
                    text-align: left;
                }}
                th {{ 
                    background-color: #2E8B57; 
                    color: white; 
                    font-weight: bold;
                }}
                tr:nth-child(even) {{ 
                    background-color: #f8f9fa; 
                }}
                .stats {{ 
                    margin-top: 20px; 
                    padding: 15px;
                    background-color: #e9f7ef;
                    border-left: 4px solid #2E8B57;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>INFORME DE DEPARTAMENTOS - ECOTECH</h1>
            </div>
            
            <div class="info">
                <p><strong>Fecha de generación:</strong> {fecha_actual}</p>
                <p><strong>Total departamentos:</strong> {len(departamentos)}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Nombre Departamento</th>
                        <th>Gerente</th>
                        <th>ID Gerente</th>
                        <th>Empleados</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for departamento in departamentos:
            num_empleados = empleados_por_depto.get(departamento.get_id(), 0)
            
            html += f"""
                    <tr>
                        <td>{departamento.get_id()}</td>
                        <td>{departamento.get_nombre()}</td>
                        <td>{departamento.get_nombre_gerente()}</td>
                        <td>{departamento.get_id_gerente()}</td>
                        <td>{num_empleados}</td>
                    </tr>
            """
        
        html += f"""
                </tbody>
            </table>
            
            <div class="stats">
                <p><strong>Estadísticas:</strong></p>
                <p>• Total empleados en departamentos: {total_empleados}</p>
                <p>• Departamentos sin empleados: {deptos_sin_empleados}</p>
            </div>
        </body>
        </html>
        """
        return html

    def _generar_reporte_departamentos_pdf_local(self, departamentos, nombre_archivo):
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            carpeta_informes = self.crear_carpeta_informes()
            ruta_pdf = os.path.join(carpeta_informes, f"{nombre_archivo}.pdf")
            
            doc = SimpleDocTemplate(ruta_pdf, pagesize=A4, topMargin=1*inch)
            elements = []
            
            styles = getSampleStyleSheet()
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1
            )
            
            titulo = Paragraph("INFORME DE DEPARTAMENTOS - ECOTECH", titulo_style)
            elements.append(titulo)
            
            fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
            info_text = f"<b>Fecha de generación:</b> {fecha_actual} | <b>Total departamentos:</b> {len(departamentos)}"
            info_paragraph = Paragraph(info_text, styles["Normal"])
            elements.append(info_paragraph)
            elements.append(Spacer(1, 0.2*inch))
            
            empleados = self.obtener_empleados()
            empleados_por_depto = {}
            for emp in empleados:
                depto_id = emp.get_id_departamento()
                if depto_id:
                    if depto_id not in empleados_por_depto:
                        empleados_por_depto[depto_id] = 0
                    empleados_por_depto[depto_id] += 1
            
            datos = [['ID', 'Nombre Departamento', 'Gerente', 'ID Gerente', 'Empleados']]
            
            for departamento in departamentos:
                num_empleados = empleados_por_depto.get(departamento.get_id(), 0)
                
                datos.append([
                    str(departamento.get_id()),
                    departamento.get_nombre(),
                    departamento.get_nombre_gerente(),
                    str(departamento.get_id_gerente()),
                    str(num_empleados)
                ])
            
            tabla = Table(datos, colWidths=[0.5*inch, 1.8*inch, 1.5*inch, 1*inch, 0.8*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
            ]))
            
            elements.append(tabla)
            
            elements.append(Spacer(1, 0.3*inch))
            total_empleados = sum(empleados_por_depto.values())
            deptos_sin_empleados = sum(1 for depto in departamentos if empleados_por_depto.get(depto.get_id(), 0) == 0)
            
            stats_text = (f"<b>Estadísticas:</b> Total empleados en departamentos: {total_empleados} | "
                        f"Departamentos sin empleados: {deptos_sin_empleados}")
            stats_paragraph = Paragraph(stats_text, styles["Normal"])
            elements.append(stats_paragraph)
            
            doc.build(elements)
            
            return ruta_pdf
            
        except Exception as e:
            print(f"Error generando PDF de departamentos: {e}")
            return f"Error: {str(e)}"
    
    def guardar_informe(self, informe):
        try:
            self.cursor = self.conn.cursor()
            sql = "INSERT INTO informes (nombre_reporte, tipo_reporte, formato, fecha_generacion, idAdministrador) VALUES (%s, %s, %s, %s, %s)"
            from datetime import datetime
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            valores = (informe.get_nombre_reporte(), informe.get_tipo_reporte(), 
                    informe.get_formato(), fecha_actual, informe.get_id_administrador())
            self.cursor.execute(sql, valores)
            self.conn.commit()
            return True
        except Exception as e:
            print(f"[DAO] Error en guardar_informe: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            return False

    def obtener_informes(self):
        try:
            self.cursor = self.conn.cursor()
            sql = "SELECT * FROM informes ORDER BY fecha_generacion DESC"
            self.cursor.execute(sql)
            resultados = self.cursor.fetchall()
            
            informes = []
            for fila in resultados:
                try:
                    from Informe import Informe
                    informe = Informe(
                        id_informe=fila[0],
                        nombre_reporte=fila[1],
                        tipo_reporte=fila[2],
                        formato=fila[3],
                        fecha_generacion=fila[4],
                        id_administrador=fila[5]
                    )
                    informes.append(informe)
                except Exception:
                    continue
            return informes
        except Exception as e:
            print(f"[DAO] Error en obtener_informes: {e}")
            return []

    def eliminar_informe(self, id_informe):
        try:
            self.cursor = self.conn.cursor()
            sql = "DELETE FROM informes WHERE idInforme = %s"
            self.cursor.execute(sql, (id_informe,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"[DAO] Error en eliminar_informe: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            return False
        
    def obtener_proyectos_por_empleado(self, id_empleado):
        try:
            self.cursor = self.conn.cursor()
            sql = """
            SELECT p.* 
            FROM proyectos p
            INNER JOIN empleados e ON p.idDepartamento = e.idDepartamento
            WHERE e.idEmpleados = %s
            """
            self.cursor.execute(sql, (id_empleado,))
            resultados = self.cursor.fetchall()
            
            proyectos = []
            for fila in resultados:
                try:
                    proyecto = Proyecto(fila[0], fila[1], fila[2], fila[3], fila[4])
                    proyectos.append(proyecto)
                except Exception:
                    continue
            return proyectos
        except Exception as e:
            print(f"[DAO] Error en obtener_proyectos_por_empleado: {e}")
            return []
        
    def asignar_empleado_proyecto(self, id_empleado, id_proyecto, rol="Miembro"):
        try:
            self.cursor = self.conn.cursor()
            sql = "INSERT INTO empleados_proyectos (idEmpleado, idProyecto, rol) VALUES (%s, %s, %s)"
            self.cursor.execute(sql, (id_empleado, id_proyecto, rol))
            self.conn.commit()
            print(f"Empleado {id_empleado} asignado al proyecto {id_proyecto}")
            return True
        except Exception as e:
            print(f"[DAO] Error en asignar_empleado_proyecto: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            return False

    def obtener_empleados_proyecto(self, id_proyecto):
        try:
            self.cursor = self.conn.cursor()
            sql = """
            SELECT e.*, ep.rol
            FROM empleados e
            INNER JOIN empleados_proyectos ep ON e.idEmpleados = ep.idEmpleado
            WHERE ep.idProyecto = %s
            ORDER BY e.nombreEmpleado ASC
            """
            self.cursor.execute(sql, (id_proyecto,))
            resultados = self.cursor.fetchall()
            
            empleados = []
            for fila in resultados:
                try:
                    empleado = EcoTech(fila[0], fila[1], fila[2], fila[3], fila[4], fila[5], fila[6], fila[7])
                    # Agregar información de la asignación
                    empleado.rol_proyecto = fila[8]  # rol
                    empleados.append(empleado)
                except Exception:
                    continue
            return empleados
        except Exception as e:
            print(f"[DAO] Error en obtener_empleados_proyecto: {e}")
            return []

    def obtener_proyectos_por_empleado(self, id_empleado):
        try:
            self.cursor = self.conn.cursor()
            sql = """
            SELECT p.*, ep.rol
            FROM proyectos p
            INNER JOIN empleados_proyectos ep ON p.idProyecto = ep.idProyecto
            WHERE ep.idEmpleado = %s
            ORDER BY p.idProyecto DESC
            """
            self.cursor.execute(sql, (id_empleado,))
            resultados = self.cursor.fetchall()
            
            proyectos = []
            for fila in resultados:
                try:
                    proyecto = Proyecto(fila[0], fila[1], fila[2], fila[3], fila[4])
                    # Agregar información de la asignación
                    proyecto.rol_empleado = fila[5]  # rol
                    proyectos.append(proyecto)
                except Exception:
                    continue
            return proyectos
        except Exception as e:
            print(f"[DAO] Error en obtener_proyectos_por_empleado: {e}")
            return []

    def desasignar_empleado_proyecto(self, id_empleado, id_proyecto):
        try:
            self.cursor = self.conn.cursor()
            sql = "DELETE FROM empleados_proyectos WHERE idEmpleado = %s AND idProyecto = %s"
            self.cursor.execute(sql, (id_empleado, id_proyecto))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"[DAO] Error en desasignar_empleado_proyecto: {e}")
            try:
                if hasattr(self, 'conn') and self.conn:
                    self.conn.rollback()
            except Exception:
                pass
            return False
        

    def obtener_escritorio(self):
        """Obtiene la ruta del escritorio del usuario"""
        sistema = platform.system()
        if sistema == "Windows":
            return os.path.join(os.path.expanduser("~"), "Desktop")
        elif sistema == "Linux":
            return os.path.join(os.path.expanduser("~"), "Escritorio")
        elif sistema == "Darwin":
            return os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            return os.path.expanduser("~")

    def crear_carpeta_informes(self):
        """Crea carpeta de informes en el escritorio si no existe"""
        escritorio = self.obtener_escritorio()
        carpeta_informes = os.path.join(escritorio, "Informes_EcoTech")
        if not os.path.exists(carpeta_informes):
            os.makedirs(carpeta_informes)
        return carpeta_informes
    
    def registrar_consulta_indicador(self, nombre, valor, fecha_ind, usuario, origen):
            try:
                self.cursor = self.conn.cursor()
                fecha_formateada = fecha_ind
                try:
                    fecha_obj = datetime.strptime(fecha_ind, "%d-%m-%Y")
                    fecha_formateada = fecha_obj.strftime("%Y-%m-%d")
                except ValueError:
                    pass
                sql = """INSERT INTO registro_indicadores 
                        (nombre_indicador, valor_indicador, fecha_indicador, fecha_consulta, usuario_consulta, origen_datos) 
                        VALUES (%s, %s, %s, NOW(), %s, %s)"""
                valores = (str(nombre), str(valor), str(fecha_formateada), str(usuario), str(origen))
                self.cursor.execute(sql, valores)
                self.conn.commit()
                print(f"Consulta de {nombre} guardada en el historial.")
                return True
            except Exception as e:
                print(f"[DAO] Error al guardar indicador: {e}")
                try:
                    if hasattr(self, 'conn') and self.conn:
                        self.conn.rollback()
                except Exception:
                    pass
                return False